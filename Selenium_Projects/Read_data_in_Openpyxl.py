import openpyxl
#Supported formats are: .xlsx,.xlsm,.xltx,.xltm
path = "E:\openpyxl_file_example.xlsx"

#To load the workbook
workbook = openpyxl.load_workbook(path)
#To select the current active sheet from workbook
sheet = workbook.active
#To select specific sheet
#sheet = workbook.get_sheet_by_name("Sheet2")
rows = sheet.max_row
cols = sheet.max_column
print("Number of rows in excel sheet:",rows)
print("Number of columns in excel sheet:",cols)

#To print the records from sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value, end="         ")
    print()